"""Workbook (XLSX) surface inventory and PHI scanning.

A workbook was previously not a recognized modality at all: an .xlsx upload was
rejected as an unsupported file. That is safe, but uninformative. It is also
not enough on its own, because callers convert workbooks to CSV by hand and
feed the result to the tabular route, which never sees the surfaces a CSV
cannot carry.

A workbook holds PHI in places a CSV has no room for: sheet names, hidden
sheets, cell comments and notes, defined names, document properties, external
links, and embedded macros. This module inventories all of them and scans every
readable text surface - including every cell of every sheet, hidden ones
included - with the same typed detector the TXT path uses.

There is no validated workbook writer, so a workbook is never releasable. The
best outcome is ``manual_review_required`` with a complete inventory. Original
bytes are never returned, and only categories and counts are reported.
"""

from __future__ import annotations

from io import BytesIO
from typing import Any, Dict, List, Optional, Tuple

from services.text_anonymization import (
    MAX_TEXT_BYTES,
    TextAnonymizationError,
    anonymize_clinical_text,
)

MAX_WORKBOOK_BYTES = 32 * 1024 * 1024
MAX_WORKBOOK_SHEETS = 100
MAX_WORKBOOK_CELLS = 200_000
# Cells are joined into batches before scanning, so the detector runs a bounded
# number of times instead of once per cell.
SCAN_BATCH_BYTES = min(32 * 1024, MAX_TEXT_BYTES)

STATUS_MANUAL_REVIEW = "manual_review_required"
STATUS_UNSCANNABLE = "unsupported_or_unscannable"

REASON_READER_UNAVAILABLE = "workbook_reader_unavailable"
REASON_UNPARSEABLE = "workbook_unparseable"
REASON_SHEET_LIMIT = "workbook_sheet_limit_exceeded"
REASON_CELL_LIMIT = "workbook_cell_limit_exceeded"
REASON_MACROS_PRESENT = "workbook_macros_present"
REASON_EXTERNAL_LINKS = "workbook_external_links_present"
REASON_EMBEDDED_OBJECTS = "workbook_embedded_objects_present"
REASON_NO_VALIDATED_WRITER = "workbook_validated_writer_unavailable"

# Distinctive markers inside the OOXML container.
_MACRO_MARKERS = (b"vbaProject.bin", b"xl/macrosheets/")
_EMBEDDED_MARKERS = (b"xl/embeddings/", b"xl/media/")
_EXTERNAL_LINK_MARKER = b"xl/externalLinks/"
_ZIP_MAGIC = b"PK\x03\x04"

_PROPERTY_FIELDS = (
    "creator",
    "title",
    "subject",
    "description",
    "keywords",
    "category",
    "lastModifiedBy",
    "company",
    "manager",
)


from services.modality_utility import (
    MEASUREMENT_VERSION,
    UNAVAILABLE,
    measure_workbook_utility,
)


class WorkbookSanitizationError(ValueError):
    def __init__(self, detail: str, status_code: int = 400):
        super().__init__(detail)
        self.detail = detail
        self.status_code = status_code


def _load_openpyxl():
    try:
        import openpyxl

        return openpyxl
    except Exception:
        return None


def _merge_counts(target: Dict[str, int], addition: Dict[str, int]) -> None:
    for key, value in addition.items():
        target[key] = target.get(key, 0) + value


def _blocked(
    reasons: List[str],
    summary: Optional[Dict[str, Any]] = None,
) -> Dict[str, Any]:
    return {
        "handler": "scan_workbook",
        "routing_status": "handler_selected",
        "anonymization_status": STATUS_UNSCANNABLE,
        "message": "Workbook could not be scanned and is not releasable.",
        "workbook_summary": summary or {},
        "unscannable_reasons": sorted(set(reasons + [REASON_NO_VALIDATED_WRITER])),
        "detected_entities": {},
        "entity_count": 0,
        "detection_sources": {},
        "utility_metrics": {
            "measurement_version": MEASUREMENT_VERSION,
            "output_available": False,
            "status": UNAVAILABLE,
            "writer_status": "no_validated_writer",
        },
        "scannable": False,
    }


class _BatchScanner:
    """Scans text in bounded batches so the detector runs a bounded number of times."""

    def __init__(self, profile: str):
        self.profile = profile
        self.counts: Dict[str, int] = {}
        self.sources: Dict[str, int] = {}
        self._pending: List[str] = []
        self._pending_bytes = 0

    def add(self, text: str) -> None:
        cleaned = str(text or "").strip()
        if not cleaned:
            return
        size = len(cleaned.encode("utf-8"))
        if size >= SCAN_BATCH_BYTES:
            # A single oversized cell is scanned on its own, truncated to the
            # batch budget rather than skipped.
            self._scan(cleaned.encode("utf-8")[:SCAN_BATCH_BYTES].decode("utf-8", "ignore"))
            return
        if self._pending_bytes + size > SCAN_BATCH_BYTES:
            self.flush()
        self._pending.append(cleaned)
        self._pending_bytes += size + 1

    def flush(self) -> None:
        if not self._pending:
            return
        batch = "\n".join(self._pending)
        self._pending = []
        self._pending_bytes = 0
        self._scan(batch)

    def _scan(self, text: str) -> None:
        if not text.strip():
            return
        result = anonymize_clinical_text(text, profile=self.profile)
        _merge_counts(self.counts, result["detected_entities"])
        _merge_counts(self.sources, result["detection_sources"])


def _document_properties(workbook) -> Tuple[List[str], List[str]]:
    """Return (present property names, their values)."""
    present: List[str] = []
    values: List[str] = []
    properties = getattr(workbook, "properties", None)
    if properties is None:
        return present, values
    for name in _PROPERTY_FIELDS:
        value = getattr(properties, name, None)
        text = str(value or "").strip()
        if text:
            present.append(name)
            values.append(text)
    return present, values


def scan_workbook_bytes(content: bytes, profile: str = "strict") -> Dict[str, Any]:
    """Inventory a workbook's PHI-bearing surfaces and scan the readable ones."""
    if not isinstance(content, (bytes, bytearray)):
        raise WorkbookSanitizationError(
            "Workbook content must be bytes", status_code=500
        )
    if not content:
        raise WorkbookSanitizationError("Uploaded workbook is empty")
    if len(content) > MAX_WORKBOOK_BYTES:
        raise WorkbookSanitizationError(
            f"Workbook uploads must be {MAX_WORKBOOK_BYTES} bytes or smaller",
            status_code=413,
        )
    raw = bytes(content)
    if not raw.startswith(_ZIP_MAGIC):
        # Legacy .xls is a compound binary, not OOXML, and is not supported.
        raise WorkbookSanitizationError("File is not an .xlsx workbook")

    openpyxl = _load_openpyxl()
    if openpyxl is None:
        return _blocked([REASON_READER_UNAVAILABLE])

    unreadable: List[str] = []
    macros_present = any(marker in raw for marker in _MACRO_MARKERS)
    embedded_present = any(marker in raw for marker in _EMBEDDED_MARKERS)
    external_links_present = _EXTERNAL_LINK_MARKER in raw
    if macros_present:
        unreadable.append(REASON_MACROS_PRESENT)
    if embedded_present:
        unreadable.append(REASON_EMBEDDED_OBJECTS)
    if external_links_present:
        unreadable.append(REASON_EXTERNAL_LINKS)

    workbook = None
    try:
        try:
            # data_only=False keeps formulas visible: a formula string can
            # itself carry an identifier or a path to one.
            workbook = openpyxl.load_workbook(
                BytesIO(raw), read_only=False, data_only=False, keep_links=False
            )
        except Exception:
            return _blocked([REASON_UNPARSEABLE] + unreadable)

        sheet_names = list(workbook.sheetnames)
        if len(sheet_names) > MAX_WORKBOOK_SHEETS:
            return _blocked(
                [REASON_SHEET_LIMIT] + unreadable, {"sheet_count": len(sheet_names)}
            )

        scanner = _BatchScanner(profile)
        hidden_sheets: List[str] = []
        comment_count = 0
        formula_cells = 0
        populated_cells = 0
        max_rows = 0
        max_columns = 0
        cell_types: Dict[str, int] = {}
        sheets: List[Dict[str, Any]] = []

        try:
            for sheet in workbook.worksheets:
                # A hidden sheet is still read and scanned; it is recorded
                # because a user may not know it is there.
                if getattr(sheet, "sheet_state", "visible") != "visible":
                    hidden_sheets.append(sheet.title)
                scanner.add(sheet.title)

                sheet_cells = 0
                sheet_comments = 0
                max_rows = max(max_rows, int(getattr(sheet, "max_row", 0) or 0))
                max_columns = max(
                    max_columns, int(getattr(sheet, "max_column", 0) or 0)
                )
                for row in sheet.iter_rows():
                    for cell in row:
                        value = cell.value
                        if value is not None and str(value).strip():
                            sheet_cells += 1
                            populated_cells += 1
                            if isinstance(value, str) and value.startswith("="):
                                formula_cells += 1
                            # The type name, never the value. A writer has to
                            # put a number back as a number, and a date back
                            # as a date, or the sheet stops computing.
                            type_name = type(value).__name__
                            cell_types[type_name] = cell_types.get(type_name, 0) + 1
                            scanner.add(str(value))
                        comment = getattr(cell, "comment", None)
                        if comment is not None and str(comment.text or "").strip():
                            sheet_comments += 1
                            comment_count += 1
                            scanner.add(str(comment.text))
                        if populated_cells > MAX_WORKBOOK_CELLS:
                            return _blocked(
                                [REASON_CELL_LIMIT] + unreadable,
                                {"sheet_count": len(sheet_names)},
                            )

                sheets.append(
                    {
                        "sheet_name": sheet.title,
                        "hidden": sheet.title in hidden_sheets,
                        "populated_cells": sheet_cells,
                        "comment_count": sheet_comments,
                    }
                )
        except Exception:
            return _blocked(
                [REASON_UNPARSEABLE] + unreadable, {"sheet_count": len(sheet_names)}
            )

        defined_names: List[str] = []
        try:
            for name in getattr(workbook, "defined_names", {}):
                defined_names.append(str(name))
                scanner.add(str(name))
        except Exception:
            unreadable.append(REASON_UNPARSEABLE)

        property_names, property_values = _document_properties(workbook)
        for value in property_values:
            scanner.add(value)

        scanner.flush()

        summary = {
            "sheet_count": len(sheet_names),
            "hidden_sheet_count": len(hidden_sheets),
            "populated_cells": populated_cells,
            "comment_count": comment_count,
            "formula_cells": formula_cells,
            "defined_name_count": len(defined_names),
            "document_properties_present": sorted(property_names),
            "macros_present": macros_present,
            "embedded_objects_present": embedded_present,
            "external_links_present": external_links_present,
            "sheets": sheets,
        }
    finally:
        if workbook is not None:
            try:
                workbook.close()
            except Exception:
                pass

    reasons = sorted(set(unreadable))
    fully_scannable = not reasons
    reasons = sorted(set(reasons + [REASON_NO_VALIDATED_WRITER]))

    return {
        "handler": "scan_workbook",
        "routing_status": "handler_selected",
        "anonymization_status": (
            STATUS_MANUAL_REVIEW if fully_scannable else STATUS_UNSCANNABLE
        ),
        "message": (
            "Workbook surfaces were inventoried and scanned. No validated "
            "workbook writer is available, so it is not automatically "
            "releasable."
        ),
        "workbook_summary": summary,
        "unscannable_reasons": reasons,
        "detected_entities": scanner.counts,
        "entity_count": sum(scanner.counts.values()),
        "detection_sources": scanner.sources,
        # Everything a workbook writer would have to preserve, counted. The
        # workbook stays under manual review; this says what "getting it
        # right" would mean, not that it has been got right.
        "utility_metrics": measure_workbook_utility(
            sheet_count=len(sheet_names),
            hidden_sheet_count=len(hidden_sheets),
            row_count=max_rows,
            column_count=max_columns,
            formula_count=formula_cells,
            comment_count=comment_count,
            defined_name_count=len(defined_names),
            document_property_count=len(property_names),
            macro_count=1 if macros_present else 0,
            external_link_count=1 if external_links_present else 0,
            cell_types=cell_types,
        ),
        "scannable": fully_scannable,
    }


def scan_workbook_for_ingestion(content: bytes, profile: str) -> Dict[str, Any]:
    """Ingestion-facing wrapper that normalizes detector failures to blocks."""
    try:
        return scan_workbook_bytes(content, profile=profile)
    except TextAnonymizationError as exc:
        raise WorkbookSanitizationError(
            exc.detail, status_code=exc.status_code
        ) from exc
